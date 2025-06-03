# начало — те же импорты и переменные
import os
import pandas as pd
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, MessageHandler, filters,
    ContextTypes, CommandHandler, CallbackQueryHandler, ConversationHandler
)
import tempfile
import logging
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import MergedCell
import zipfile

logging.basicConfig(format='[LOG] %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

TEMPLATE_FILENAME = "AllPackageEC_.xlsx"
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), TEMPLATE_FILENAME)

MODE_CHOICE = {}
LOAD_SOURCE, LOAD_PINFL = range(2)
USER_FILES = {}

def get_main_keyboard():
    keyboard = [
        [InlineKeyboardButton("▶️ Обработать на части (1000)", callback_data="chunk")],
        [InlineKeyboardButton("▶️ Обработать на части (500)", callback_data="chunk500")],
        [InlineKeyboardButton("▶️ Обработать на части (250)", callback_data="chunk250")],
        [InlineKeyboardButton("📄 Макрос Пасспорт", callback_data="passport")],
        [InlineKeyboardButton("🔄 Замена ПИНФЛ", callback_data="replace_pinfl")]
    ]
    return InlineKeyboardMarkup(keyboard)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Выберите режим обработки:", reply_markup=get_main_keyboard())

async def mode_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    MODE_CHOICE[user_id] = query.data

    if query.data == "replace_pinfl":
        await query.message.reply_text("Пожалуйста, загрузите файл реестра (source_file).")
        return LOAD_SOURCE

    await query.message.reply_text("Отправьте Excel-файл для выбранной обработки.")
    return ConversationHandler.END

async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.message.from_user
    user_id = user.id
    mode = MODE_CHOICE.get(user_id)

    if not mode:
        await update.message.reply_text("Сначала выберите режим через /start.")
        return

    document = update.message.document
    file = await context.bot.get_file(document.file_id)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
        await file.download_to_drive(f.name)
        data_file = f.name

    if mode == "chunk":
        await process_in_parts(update, context, data_file, chunk_size=1000)
    elif mode == "chunk500":
        await process_in_parts(update, context, data_file, chunk_size=500)
    elif mode == "chunk250":
        await process_in_parts(update, context, data_file, chunk_size=250, dynamic_naming=True)
    elif mode == "passport":
        await process_passport_macro(update, context, data_file)

    await update.message.reply_text("Выберите режим обработки:", reply_markup=get_main_keyboard())

async def process_in_parts(update, context, data_file, chunk_size=1000, dynamic_naming=False):
    logger.info(f"Обработка: разбивка на части по {chunk_size} шт.")
    df = pd.read_excel(data_file, header=None, skiprows=3)

    def fix_code(x):
        try:
            s = str(int(float(x)))
            if len(s) == 5:
                return "0" + s
            return x
        except:
            return x

    df[10] = df[10].apply(fix_code)

    seen = set()
    for idx, val in df[0].items():
        val = str(val).strip()
        if val and val in seen:
            df.loc[idx, 0:7] = None
        else:
            seen.add(val)

    parts = [df[i:i + chunk_size] for i in range(0, len(df), chunk_size)]

    output_files = []
    for idx, part in enumerate(parts):
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        for r_idx, row in enumerate(dataframe_to_rows(part, index=False, header=False), start=4):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        start_index = idx * chunk_size if not dynamic_naming else idx * 250
        filename = f"AllPackageEC_{start_index}.xlsx"
        output_path = os.path.join(tempfile.gettempdir(), filename)
        wb.save(output_path)
        output_files.append(output_path)
        logger.info(f"Сохранён файл: {filename}")

    zip_path = os.path.join(tempfile.gettempdir(), f"AllPackageEC_{update.message.from_user.username}.zip")
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        for file_path in output_files:
            zipf.write(file_path, os.path.basename(file_path))

    await update.message.reply_text("Обработка завершена. Архив отправляется...")
    await context.bot.send_document(chat_id=update.message.chat_id, document=open(zip_path, 'rb'))

async def process_passport_macro(update, context, data_file):
    logger.info("Выполняется макрос 'Пасспорт'")
    wb = load_workbook(data_file)
    ws = wb.active

    valid_start = "123456789MRTGKZECUVFBNDGHJLKQIP"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        val = str(row[4].value).strip() if row[4].value else ""
        if val and val[0].upper() in valid_start:
            row[4].value = "AB0663236"
            row[5].value = "23,12,1988"

    output_path = os.path.join(tempfile.gettempdir(), f"PassportUpdated_{update.message.from_user.username}.xlsx")
    wb.save(output_path)

    await update.message.reply_text("Макрос выполнен. Файл отправляется...")
    await context.bot.send_document(chat_id=update.message.chat_id, document=open(output_path, 'rb'))

# --- ПИНФЛ замена ---
async def load_source_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    document = update.message.document
    if not document or not document.file_name.endswith('.xlsx'):
        await update.message.reply_text("Пожалуйста, загрузите корректный Excel-файл реестра.")
        return LOAD_SOURCE

    file = await context.bot.get_file(document.file_id)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
        await file.download_to_drive(f.name)
        USER_FILES[update.message.from_user.id] = {'source_file': f.name}

    await update.message.reply_text("Файл реестра получен. Теперь загрузите файл с результатами ПИНФЛ.")
    return LOAD_PINFL

async def load_pinfl_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    document = update.message.document
    if not document or not document.file_name.endswith('.xlsx'):
        await update.message.reply_text("Пожалуйста, загрузите корректный Excel-файл результатов ПИНФЛ.")
        return LOAD_PINFL

    file = await context.bot.get_file(document.file_id)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
        await file.download_to_drive(f.name)
        USER_FILES[update.message.from_user.id]['pinfl_file'] = f.name

    user_id = update.message.from_user.id
    files = USER_FILES.get(user_id)
    if not files or 'source_file' not in files or 'pinfl_file' not in files:
        await update.message.reply_text("Ошибка: не все файлы загружены.")
        return ConversationHandler.END

    await update.message.reply_text("Файлы получены. Выполняется замена ПИНФЛ...")

    output_file = os.path.join(tempfile.gettempdir(), f"AllPackageEC_GOOD_{user_id}.xlsx")
    try:
        replace_pinfl(files['source_file'], files['pinfl_file'], output_file)
    except Exception as e:
        await update.message.reply_text(f"Ошибка при замене ПИНФЛ: {e}")
        return ConversationHandler.END

    await update.message.reply_text("Замена ПИНФЛ завершена. Отправляю файл...")
    await context.bot.send_document(chat_id=update.message.chat_id, document=open(output_file, 'rb'))

    USER_FILES.pop(user_id, None)
    await update.message.reply_text("Выберите режим обработки:", reply_markup=get_main_keyboard())
    return ConversationHandler.END

def replace_pinfl(source_file, pinfl_file, output_file):
    df2 = pd.read_excel(pinfl_file, header=None)
    passport_to_pinfl = dict(
        zip(df2.iloc[:, 8].astype(str).str.strip().str.upper(), df2.iloc[:, 9])
    )

    wb = load_workbook(filename=source_file)
    ws = wb.active

    valid_start = tuple('0123456789KJTIFHBMNCXZSDQWRYUPLE')
    replacements = []

    for row in ws.iter_rows(min_row=2):
        cell_e = row[4]
        cell_f = row[5]
        val = cell_e.value

        if val is None or str(val).strip() == '':
            cell_e.value = 'AB0663236'
            if not isinstance(cell_f, MergedCell):
                cell_f.value = '23.12.1988'
            continue

        key = str(val).strip().upper()
        if key.startswith(valid_start):
            pinfl = passport_to_pinfl.get(key)
            if pinfl:
                replacements.append((val, pinfl))
                cell_e.value = pinfl

    wb.save(output_file)
    with open('замены_log.txt', 'w', encoding='utf-8') as log_file:
        for old, new in replacements:
            log_file.write(f'{old} → {new}\n')

    logger.info(f'Готово! Файл сохранён как {output_file}')
    logger.info(f'Заменено {len(replacements)} паспортов.')

def main():
    app = ApplicationBuilder().token("7872241701:AAF633V3rjyXTJkD8F0lEW13nDtAqHoqeic").build()

    conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(mode_selected)],
        states={
            LOAD_SOURCE: [MessageHandler(filters.Document.FileExtension("xlsx"), load_source_file)],
            LOAD_PINFL: [MessageHandler(filters.Document.FileExtension("xlsx"), load_pinfl_file)],
        },
        fallbacks=[],
        allow_reentry=True,
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(conv_handler)
    app.add_handler(MessageHandler(filters.Document.FileExtension("xlsx"), handle_file))

    logger.info("Бот запущен")
    app.run_polling()

if __name__ == "__main__":
    main()
